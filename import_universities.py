"""
Import university sourcing data from the XLSX master database.

Reads three sheets:
  - "University Data"       → name, country, city, rank, score, spinouts, funding, labs, clubs, priority
  - "Scraping Feasibility"  → sourcing_url, secondary_urls, page_type, scraping_score, update_frequency
  - "Contact Directory"     → primary_contact_email

Joins on rank + university name and inserts into the universities table.
"""

import sys
import os

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from database.database import init_db, get_connection

XLSX_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "data",
    "University_Sourcing_Master_Database.xlsx",
)


def read_sheet_as_dicts(wb, sheet_name):
    """Read a sheet into a list of dicts keyed by header row."""
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        rows.append(d)
    return rows


def build_lookup(rows, key="Rank"):
    """Build a dict keyed by rank for fast joining."""
    return {r[key]: r for r in rows if r.get(key) is not None}


def clean(val):
    """Normalize a cell value: strip strings, convert None."""
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        return val if val else None
    return val


def to_int(val):
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def to_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def main():
    init_db()

    print("University Importer")
    print("=" * 50)

    if not os.path.exists(XLSX_PATH):
        print(f"ERROR: File not found: {XLSX_PATH}")
        sys.exit(1)

    print(f"\nReading {XLSX_PATH}...")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    # Read the three data sheets
    uni_rows = read_sheet_as_dicts(wb, "University Data")
    scrape_rows = read_sheet_as_dicts(wb, "Scraping Feasibility")
    contact_rows = read_sheet_as_dicts(wb, "Contact Directory")

    # Build lookups by rank
    scrape_by_rank = build_lookup(scrape_rows)
    contact_by_rank = build_lookup(contact_rows)

    print(f"  University Data:      {len(uni_rows)} rows")
    print(f"  Scraping Feasibility: {len(scrape_rows)} rows")
    print(f"  Contact Directory:    {len(contact_rows)} rows")

    # Clear existing data to allow re-runs
    conn = get_connection()
    existing = conn.execute("SELECT COUNT(*) FROM universities").fetchone()[0]
    if existing:
        conn.execute("DELETE FROM universities")
        conn.commit()
        print(f"\n  Cleared {existing} existing rows")
    conn.close()

    # Import
    imported = 0
    countries = {}

    for row in uni_rows:
        rank = row.get("Rank")
        name = clean(row.get("University"))
        if not name:
            continue

        country = clean(row.get("Country"))
        city = clean(row.get("City"))
        score = to_int(row.get("Score"))
        num_spinouts = to_int(row.get("Number of VC-backed spinouts"))
        total_funding = clean(row.get("Total funding raised"))
        ai_labs = clean(row.get("AI Labs - Names"))
        clubs = clean(row.get("Clubs - Names"))
        scout_priority = to_float(row.get("Scout Priority Score"))

        # Join scraping feasibility data
        scrape = scrape_by_rank.get(rank, {})
        sourcing_url = clean(scrape.get("Sourcing Link"))
        secondary_urls = clean(scrape.get("Sourcing Link - Secondary URLs"))
        page_type = clean(scrape.get("Sourcing Link - Page Type"))
        scraping_score = to_float(scrape.get("Sourcing Link - Scraping Score"))
        update_frequency = clean(scrape.get("Sourcing Link - Update Frequency"))

        # Join contact data
        contact = contact_by_rank.get(rank, {})
        primary_email = clean(contact.get("Primary Contact Email"))

        conn = get_connection()
        conn.execute(
            """INSERT INTO universities
               (name, country, city, rank, score, num_spinouts, total_funding,
                sourcing_url, secondary_urls, page_type, scraping_score,
                update_frequency, scout_priority_score, ai_labs, clubs,
                primary_contact_email)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (name, country, city, to_int(rank), score, num_spinouts, total_funding,
             sourcing_url, secondary_urls, page_type, scraping_score,
             update_frequency, scout_priority, ai_labs, clubs, primary_email),
        )
        conn.commit()
        conn.close()

        imported += 1
        countries[country] = countries.get(country, 0) + 1

    # Summary
    print(f"\n  Imported {imported} universities")

    print(f"\n  By country:")
    for c, n in sorted(countries.items(), key=lambda x: -x[1]):
        print(f"    {c or 'Unknown':30s} {n:>3}")

    # Quick stats
    conn = get_connection()
    with_url = conn.execute(
        "SELECT COUNT(*) FROM universities WHERE sourcing_url IS NOT NULL"
    ).fetchone()[0]
    with_labs = conn.execute(
        "SELECT COUNT(*) FROM universities WHERE ai_labs IS NOT NULL"
    ).fetchone()[0]
    with_clubs = conn.execute(
        "SELECT COUNT(*) FROM universities WHERE clubs IS NOT NULL"
    ).fetchone()[0]
    with_email = conn.execute(
        "SELECT COUNT(*) FROM universities WHERE primary_contact_email IS NOT NULL"
    ).fetchone()[0]
    avg_scrape = conn.execute(
        "SELECT AVG(scraping_score) FROM universities WHERE scraping_score IS NOT NULL"
    ).fetchone()[0]
    top5 = conn.execute(
        "SELECT rank, name, country, score, num_spinouts, total_funding "
        "FROM universities ORDER BY rank LIMIT 5"
    ).fetchall()
    conn.close()

    print(f"\n  Coverage:")
    print(f"    With sourcing URL:       {with_url:>3}/{imported}")
    print(f"    With AI labs:            {with_labs:>3}/{imported}")
    print(f"    With clubs:              {with_clubs:>3}/{imported}")
    print(f"    With contact email:      {with_email:>3}/{imported}")
    print(f"    Avg scraping score:      {avg_scrape:.1f}/5" if avg_scrape else "")

    print(f"\n  Top 5 universities:")
    print(f"    {'Rank':>4}  {'University':40s}  {'Country':20s}  {'Score':>5}  {'Spinouts':>8}  {'Funding':>10}")
    print(f"    {'─'*4}  {'─'*40}  {'─'*20}  {'─'*5}  {'─'*8}  {'─'*10}")
    for r in top5:
        print(f"    {r[0]:>4}  {r[1]:40s}  {r[2]:20s}  {r[3]:>5}  {r[4]:>8}  {r[5]:>10}")

    print()


if __name__ == "__main__":
    main()
